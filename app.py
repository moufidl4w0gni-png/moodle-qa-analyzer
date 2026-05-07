"""
Analyseur & Correcteur de questions Moodle
==========================================
Fonctionnalités :
- Import XML Moodle et GIFT
- Détection avancée des erreurs (HTML, LaTeX, graphiques, réponses)
- Correction automatique du barème Moodle (fraction 100 / -33.333)
- Export XML corrigé + rapport Excel
- Vérification optionnelle par IA (Claude API)
"""

import xml.etree.ElementTree as ET
import re
import io
import copy
import html
from pathlib import Path

import pandas as pd
import streamlit as st
import requests

# ─────────────────────────────────────────────────────────────
# CONFIG PAGE
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Analyseur Moodle QA",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────────────────────
# CLÉ API (optionnelle)
# ─────────────────────────────────────────────────────────────
ANTHROPIC_API_KEY = st.secrets.get("ANTHROPIC_API_KEY", "") if hasattr(st, "secrets") else ""

# ─────────────────────────────────────────────────────────────
# CSS CUSTOM
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main .block-container { padding-top: 1.5rem; }
    .stDataFrame { border-radius: 8px; }
    .metric-card {
        background: #f8f9fa;
        border: 1px solid #e9ecef;
        border-radius: 8px;
        padding: 12px 16px;
        text-align: center;
    }
    .badge-ok    { color: #155724; background:#d4edda; padding:2px 8px; border-radius:12px; font-size:12px; }
    .badge-warn  { color: #856404; background:#fff3cd; padding:2px 8px; border-radius:12px; font-size:12px; }
    .badge-err   { color: #721c24; background:#f8d7da; padding:2px 8px; border-radius:12px; font-size:12px; }
    .badge-fixed { color: #0c5460; background:#d1ecf1; padding:2px 8px; border-radius:12px; font-size:12px; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# 1. PARSERS
# ═══════════════════════════════════════════════════════════════

def get_text(element):
    """Extrait le texte d'un élément XML Moodle (gère CDATA, HTML embarqué, itertext)."""
    if element is None:
        return ""
    # itertext() récupère tout le contenu textuel y compris CDATA et enfants imbriqués
    parts = list(element.itertext())
    text = "".join(parts).strip()
    return text


def parse_fraction(raw):
    """Convertit une fraction Moodle en float, sans crash."""
    try:
        return float((raw or "0").strip())
    except (ValueError, AttributeError):
        return 0.0


def charger_xml(buffer):
    """Parse un fichier XML Moodle et retourne la liste des questions."""
    questions = []
    try:
        buffer.seek(0)
        content = buffer.read()
        # Nettoyage des caractères illégaux XML
        content = re.sub(rb'[\x00-\x08\x0b\x0c\x0e-\x1f]', b'', content)
        root = ET.fromstring(content)
    except ET.ParseError as e:
        st.error(f"❌ Erreur de parsing XML : {e}")
        return []

    for q in root.findall("question"):
        qtype = q.get("type", "")
        if qtype in ("category", "description", ""):
            continue

        nom   = get_text(q.find("name/text")) or "(sans nom)"
        # Extraction robuste du texte de la question (CDATA, HTML, formats variés)
        qt_elem = q.find("questiontext")
        texte = ""
        if qt_elem is not None:
            text_elem = qt_elem.find("text")
            if text_elem is not None:
                texte = get_text(text_elem)
            if not texte:
                # fallback : tout le contenu de questiontext
                texte = get_text(qt_elem)
        if not texte:
            # dernier fallback : cherche un <text> direct
            texte = get_text(q.find("questiontext/text")) or ""

        # Pénalité & grade par défaut
        defaultgrade = parse_fraction(q.findtext("defaultgrade"))
        penalty      = parse_fraction(q.findtext("penalty"))
        shuffle      = q.findtext("shuffleanswers", "true")

        reponses = []
        best_frac = -999
        bonne = None

        for ans in q.findall("answer"):
            txt  = get_text(ans.find("text")) or ans.findtext("text", "")
            frac = parse_fraction(ans.get("fraction"))
            fb   = get_text(ans.find("feedback/text"))
            reponses.append({"texte": txt.strip(), "fraction": frac, "feedback": fb})
            if frac > best_frac:
                best_frac = frac
                bonne = txt.strip()

        if best_frac <= 0:
            bonne = None

        questions.append({
            "id":           nom,
            "nom":          nom,
            "type":         qtype,
            "texte":        texte,
            "reponses":     reponses,
            "bonne":        bonne,
            "best_frac":    best_frac,
            "defaultgrade": defaultgrade if defaultgrade > 0 else 1.0,
            "penalty":      penalty,
            "shuffle":      shuffle,
            "raw":          q,  # élément XML original pour correction
        })
    return questions


def charger_gift(buffer):
    """Parse un fichier GIFT et retourne la liste des questions."""
    questions = []
    try:
        buffer.seek(0)
        try:
            contenu = buffer.read().decode("utf-8")
        except UnicodeDecodeError:
            buffer.seek(0)
            contenu = buffer.read().decode("latin-1")
    except Exception as e:
        st.error(f"❌ Erreur lecture GIFT : {e}")
        return []

    contenu = contenu.replace("\r\n", "\n").replace("\r", "\n")
    # Supprime commentaires //
    lines = [l for l in contenu.split("\n") if not l.strip().startswith("//")]
    contenu = "\n".join(lines)

    blocs = contenu.split("::")
    for bloc in blocs[1:]:
        if "::" not in bloc or "{" not in bloc:
            continue
        try:
            nom, reste = bloc.split("::", 1)
        except ValueError:
            continue

        nom = nom.strip()
        m = re.search(r'\{([\s\S]+?)\}', reste)
        if not m:
            continue

        texte = reste[:m.start()].strip()
        reponses_brut = m.group(1)
        reponses = []
        bonne = None

        for line in reponses_brut.split("\n"):
            line = line.strip()
            if not line:
                continue
            # Fraction partielle ~%50%...
            m_frac = re.match(r'~%(-?\d+(?:\.\d+)?)%(.+)', line)
            if line.startswith("="):
                txt = line[1:].split("#")[0].strip()
                bonne = txt
                reponses.append({"texte": txt, "fraction": 100.0, "feedback": ""})
            elif m_frac:
                frac = float(m_frac.group(1))
                txt  = m_frac.group(2).split("#")[0].strip()
                reponses.append({"texte": txt, "fraction": frac, "feedback": ""})
            elif line.startswith("~"):
                txt = line[1:].split("#")[0].strip()
                reponses.append({"texte": txt, "fraction": 0.0, "feedback": ""})

        if not reponses:
            continue

        questions.append({
            "id":           nom,
            "nom":          nom,
            "type":         "multichoice",
            "texte":        texte,
            "reponses":     reponses,
            "bonne":        bonne,
            "best_frac":    100.0 if bonne else -1,
            "defaultgrade": 1.0,
            "penalty":      0.333333,
            "shuffle":      "true",
            "raw":          None,
        })
    return questions


# ═══════════════════════════════════════════════════════════════
# 2. DÉTECTION D'ERREURS (avancée)
# ═══════════════════════════════════════════════════════════════

def check_html_latex(texte):
    """
    Vérifie :
    - Balises HTML ouvertes non fermées (heuristique avancée)
    - LaTeX $$ impair
    - LaTeX \\( \\) déséquilibré
    - LaTeX \\[ \\] déséquilibré
    - Entités HTML invalides
    """
    issues = []

    # Balises HTML — ignore les balises auto-fermantes et les vides
    open_tags  = re.findall(r'<([a-zA-Z][a-zA-Z0-9]*)\b[^>]*(?<!/)>', texte)
    close_tags = re.findall(r'</([a-zA-Z][a-zA-Z0-9]*)>', texte)
    self_close = re.findall(r'<[a-zA-Z][^>]*/>', texte)
    void_tags  = {'br','hr','img','input','link','meta','area','base','col','embed','source','track','wbr'}
    open_tags  = [t.lower() for t in open_tags if t.lower() not in void_tags]
    close_tags = [t.lower() for t in close_tags]
    if sorted(open_tags) != sorted(close_tags):
        issues.append(f"Balises HTML déséquilibrées (ouv:{len(open_tags)} ferm:{len(close_tags)})")

    # LaTeX $$
    if texte.count("$$") % 2 != 0:
        issues.append("LaTeX $$ non fermé")

    # LaTeX \( \)
    if texte.count("\\(") != texte.count("\\)"):
        issues.append("LaTeX \\( \\) déséquilibré")

    # LaTeX \[ \]
    if texte.count("\\[") != texte.count("\\]"):
        issues.append("LaTeX \\[ \\] déséquilibré")

    # Entités HTML invalides (ex: & sans ;)
    raw_amp = re.findall(r'&(?!(?:[a-zA-Z]+|#\d+|#x[0-9a-fA-F]+);)', texte)
    if raw_amp:
        issues.append(f"Entité HTML invalide (&...;) : {len(raw_amp)} occurrence(s)")

    # Balises <img> sans attribut src
    for img in re.finditer(r'<img([^>]*)>', texte, re.IGNORECASE):
        if 'src' not in img.group(1):
            issues.append("Balise <img> sans attribut src")
            break

    if not issues:
        return "✅ OK", "ok"
    return " | ".join(issues), "err" if any("déséquili" in i or "invalide" in i for i in issues) else "warn"


def check_graphique(texte):
    """Détecte les références visuelles sans image associée."""
    mots = [
        "graphique", "courbe", "figure", "ci-dessous", "diagramme",
        "représentation", "schéma", "tableau ci", "illustration",
        "voir ci", "image ci", "tracé", "nuage de points"
    ]
    texte_lower = texte.lower()
    if any(m in texte_lower for m in mots):
        if "<img" not in texte and "![" not in texte:
            return "⚠️ Référence visuelle sans image", "warn"
    return "✅ OK", "ok"


def check_reponses(reponses, bonne, qtype):
    """Vérifie la cohérence des réponses selon le type."""
    issues = []

    if not reponses:
        return "❌ Aucune réponse définie", "err"

    if qtype in ("multichoice", "multichoiceset"):
        if bonne is None:
            issues.append("Pas de bonne réponse définie")
        if len(reponses) < 2:
            issues.append(f"Seulement {len(reponses)} réponse(s) proposée(s)")
        if len(reponses) > 6:
            issues.append(f"Beaucoup de réponses ({len(reponses)}) — vérifier")

        fracs = [r["fraction"] for r in reponses]
        if max(fracs) <= 0:
            issues.append("Aucune fraction positive (pas de bonne réponse)")

        # Vérifie doublons
        textes = [r["texte"].strip().lower() for r in reponses if r["texte"]]
        if len(textes) != len(set(textes)):
            issues.append("Réponses dupliquées détectées")

        # Vérifie réponses vides
        vides = [r for r in reponses if not r["texte"].strip()]
        if vides:
            issues.append(f"{len(vides)} réponse(s) vide(s)")

    elif qtype == "truefalse":
        if len(reponses) != 2:
            issues.append(f"Question Vrai/Faux doit avoir 2 réponses, {len(reponses)} trouvée(s)")

    if not issues:
        return "✅ OK", "ok"
    sev = "err" if any(k in " ".join(issues) for k in ["pas de bonne", "aucune fraction", "aucune réponse"]) else "warn"
    return " | ".join(issues), sev


def check_bareme(reponses, penalty, best_frac, qtype):
    """
    Vérifie si le barème est conforme aux standards Moodle.
    Gère les deux modes Moodle :
    - QCM une seule bonne réponse  : fraction = 100%
    - QCM plusieurs bonnes réponses : fractions partielles dont la somme = 100%
      (ex: 2 bonnes réponses → 50%+50%, 3 → 33.33%+33.33%+33.33%)
    """
    issues = []
    corrections = []

    if qtype not in ("multichoice", "multichoiceset", "truefalse"):
        return "✅ N/A", "ok", []

    fracs = [r["fraction"] for r in reponses]
    pos_fracs = [f for f in fracs if f > 0]
    somme_pos = round(sum(pos_fracs), 2) if pos_fracs else 0.0
    nb_bonnes = len(pos_fracs)

    # ── Détection du mode : une seule bonne réponse ou plusieurs ──
    # Mode multi-réponses : plusieurs fracs positives dont la somme = 100%
    is_multi_answer = nb_bonnes > 1 and abs(somme_pos - 100.0) < 0.5

    if not is_multi_answer:
        # Mode réponse unique : la meilleure fraction doit être 100%
        if best_frac > 0 and abs(best_frac - 100.0) > 0.5:
            issues.append(f"Bonne réponse à {best_frac:.5g}% au lieu de 100%")
            corrections.append(("best_fraction", best_frac, 100.0))

        # Somme des fractions positives doit être 100
        if pos_fracs and abs(somme_pos - 100.0) > 0.5:
            issues.append(f"Somme des fractions positives = {somme_pos:.1f}% ≠ 100%")
    else:
        # Mode multi-réponses : valider que chaque fraction partielle est cohérente
        frac_attendue = round(100.0 / nb_bonnes, 5)
        for i, r in enumerate(reponses):
            f = r["fraction"]
            if f > 0 and abs(f - frac_attendue) > 1.0:
                issues.append(
                    f"Fraction partielle {f:.5g}% inhabituelle pour {nb_bonnes} bonnes réponses "                    f"(attendu ≈{frac_attendue:.5g}%)")
                break  # signaler une seule fois

    # Fractions invalides (> 100 ou < -100) — toujours vérifier
    for i, f in enumerate(fracs):
        if f > 100.5:
            issues.append(f"Fraction {f:.5g}% > 100% (rép. {i+1})")
            corrections.append(("fraction_over", i, 100.0))
        if f < -100.5:
            issues.append(f"Fraction {f:.5g}% < -100% (rép. {i+1})")
            corrections.append(("fraction_under", i, -100.0))

    # Pénalité non standard
    valid_penalties = {0.0, 0.1, 0.3333333, 0.5, 1.0}
    if round(penalty, 4) not in {round(v, 4) for v in valid_penalties}:
        issues.append(f"Pénalité {penalty:.4f} non standard (attendu: 0, 0.1, 0.3333, 0.5 ou 1)")
        corrections.append(("penalty", penalty, 0.3333333))

    if not issues:
        mode = f" ({nb_bonnes} bonnes réponses × {best_frac:.5g}%)" if is_multi_answer else ""
        return f"✅ Correct{mode}", "ok", []
    sev = "err" if corrections else "warn"
    return " | ".join(issues), sev, corrections


def check_texte_vide(texte):
    """Vérification du texte désactivée — toujours OK."""
    return "✅ OK", "ok"


def check_encodage(texte):
    """Détecte les problèmes d'encodage courants."""
    bad_patterns = [
        (r'Ã©', 'é mal encodé'),
        (r'Ã\xa0', 'à mal encodé'),
        (r'Ã¨', 'è mal encodé'),
        (r'â€™', "apostrophe mal encodée"),
        (r'â€œ|â€\x9d', "guillemets mal encodés"),
        (r'\?\?+', "caractères indéfinis (??)"),
    ]
    found = []
    for pattern, label in bad_patterns:
        if re.search(pattern, texte):
            found.append(label)
    if found:
        return "⚠️ " + " | ".join(found), "warn"
    return "✅ OK", "ok"


def analyser_question(q):
    """Analyse complète d'une question, retourne un dict de résultats."""
    texte = q["texte"]
    chk_html,   sev_html   = check_html_latex(texte)
    chk_graph,  sev_graph  = check_graphique(texte)
    chk_rep,    sev_rep    = check_reponses(q["reponses"], q["bonne"], q["type"])
    chk_texte,  sev_texte  = check_texte_vide(texte)
    chk_enc,    sev_enc    = check_encodage(texte)
    chk_bar,    sev_bar, corr_bar = check_bareme(
        q["reponses"], q["penalty"], q["best_frac"], q["type"]
    )

    severities = [sev_html, sev_graph, sev_rep, sev_enc, sev_bar]
    global_sev = "err" if "err" in severities else "warn" if "warn" in severities else "ok"

    return {
        "Nom":             q.get("nom_affiche", q["nom"]),
        "Type":            q["type"],
        "Nb réponses":     len(q["reponses"]),
        "Bonne réponse":   (q["bonne"] or "—")[:60],
        "HTML / LaTeX":    chk_html,
        "Graphique":       chk_graph,
        "Réponses":        chk_rep,
        "Encodage":        chk_enc,
        "Barème":          chk_bar,
        "Statut global":   "❌ Erreur" if global_sev == "err" else "⚠️ Avertissement" if global_sev == "warn" else "✅ OK",
        "_corrections_bareme": corr_bar,
        "_q": q,
    }


# ═══════════════════════════════════════════════════════════════
# 3. CORRECTION AUTOMATIQUE DU BARÈME
# ═══════════════════════════════════════════════════════════════

def corriger_bareme_question(q, penalty_cible=0.3333333):
    """
    Corrige automatiquement le barème d'une question :
    - Bonne réponse → fraction = 100
    - Mauvaises réponses → fraction = 0 (ou −33.3333 si pénalité activée dans les réponses)
    - Pénalité → penalty_cible
    - defaultgrade inchangé
    Retourne la question corrigée et un log des changements.
    """
    q_corr = copy.deepcopy(q)
    log = []

    # Trouver la meilleure fraction actuelle
    if not q_corr["reponses"]:
        return q_corr, log

    fracs = [r["fraction"] for r in q_corr["reponses"]]
    max_frac = max(fracs)

    for r in q_corr["reponses"]:
        old = r["fraction"]
        if r["fraction"] == max_frac and max_frac > 0:
            # C'est la bonne réponse
            if r["fraction"] != 100.0:
                r["fraction"] = 100.0
                log.append(f"Bonne réponse '{r['texte'][:30]}' : {old}% → 100%")
        else:
            # Mauvaise réponse : doit être 0 ou négatif selon la politique
            if r["fraction"] > 0:
                r["fraction"] = 0.0
                log.append(f"Mauvaise réponse '{r['texte'][:30]}' : {old}% → 0%")
            elif r["fraction"] < -100:
                r["fraction"] = -100.0
                log.append(f"Fraction invalide '{r['texte'][:30]}' : {old}% → -100%")

    # Pénalité
    if round(q_corr["penalty"], 4) != round(penalty_cible, 4):
        log.append(f"Pénalité : {q_corr['penalty']:.4f} → {penalty_cible:.7f}")
        q_corr["penalty"] = penalty_cible

    return q_corr, log


def corriger_toutes_questions(questions, penalty_cible=0.3333333):
    """Corrige le barème de toutes les questions multichoice."""
    questions_corrigees = []
    tous_logs = {}
    for q in questions:
        if q["type"] in ("multichoice", "multichoiceset", "truefalse"):
            q_c, log = corriger_bareme_question(q, penalty_cible)
            questions_corrigees.append(q_c)
            if log:
                tous_logs[q["nom"]] = log
        else:
            questions_corrigees.append(q)
    return questions_corrigees, tous_logs


# ═══════════════════════════════════════════════════════════════
# 4. EXPORT XML MOODLE CORRIGÉ
# ═══════════════════════════════════════════════════════════════

def escape_xml(text):
    return html.escape(str(text), quote=True)


def generer_xml_corrige(questions):
    """Génère un fichier XML Moodle valide à partir des questions corrigées."""
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', '<quiz>', '']

    for q in questions:
        lines.append(f'  <question type="{escape_xml(q["type"])}">')
        lines.append(f'    <name><text>{escape_xml(q["nom"])}</text></name>')
        lines.append(f'    <questiontext format="html">')
        lines.append(f'      <text><![CDATA[{q["texte"]}]]></text>')
        lines.append(f'    </questiontext>')
        lines.append(f'    <defaultgrade>{q["defaultgrade"]}</defaultgrade>')
        lines.append(f'    <penalty>{q["penalty"]:.7f}</penalty>')
        lines.append(f'    <hidden>0</hidden>')
        lines.append(f'    <single>true</single>')
        lines.append(f'    <shuffleanswers>{q["shuffle"]}</shuffleanswers>')
        lines.append(f'    <answernumbering>abc</answernumbering>')

        for r in q["reponses"]:
            frac = r["fraction"]
            frac_str = f'{frac:.5f}' if frac != int(frac) else str(int(frac))
            lines.append(f'    <answer fraction="{frac_str}">')
            lines.append(f'      <text><![CDATA[{r["texte"]}]]></text>')
            if r.get("feedback"):
                lines.append(f'      <feedback><text><![CDATA[{r["feedback"]}]]></text></feedback>')
            else:
                lines.append(f'      <feedback><text></text></feedback>')
            lines.append(f'    </answer>')

        lines.append(f'  </question>')
        lines.append('')

    lines.append('</quiz>')
    return '\n'.join(lines)


# ═══════════════════════════════════════════════════════════════
# 5. VÉRIFICATION IA (Claude)
# ═══════════════════════════════════════════════════════════════

def verifier_ia(texte, reponses, bonne):
    """Vérifie la cohérence d'une question via Claude API."""
    if not ANTHROPIC_API_KEY:
        return "API non configurée"
    if not bonne:
        return "Pas de bonne réponse à vérifier"

    rep_str = "\n".join(
        f"{'[✓] ' if r['texte'] == bonne else '[ ] '}{r['texte']}"
        for r in reponses[:6]
    )
    prompt = (
        f"Question d'examen :\n{texte}\n\n"
        f"Réponses (✓ = indiquée comme bonne) :\n{rep_str}\n\n"
        "La bonne réponse indiquée est-elle correcte ? "
        "Réponds uniquement par OUI, NON ou INCERTAIN, suivi d'une justification d'une phrase max."
    )
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type": "application/json",
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01"
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 120,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=20
        )
        data = r.json()
        if "content" in data and data["content"]:
            return data["content"][0]["text"].strip()
        return f"Erreur API : {data.get('error', {}).get('message', 'inconnue')}"
    except Exception as e:
        return f"Erreur : {str(e)[:60]}"


# ═══════════════════════════════════════════════════════════════
# 6. EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════

def generer_excel(df_analyse, logs_correction):
    """Génère un rapport Excel multi-feuilles."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Feuille 1 — Analyse complète
        cols_export = [c for c in df_analyse.columns if not c.startswith("_")]
        df_analyse[cols_export].to_excel(writer, index=False, sheet_name="Analyse complète")

        # Feuille 2 — Problèmes seulement
        mask = df_analyse["Statut global"] != "✅ OK"
        if mask.any():
            df_analyse[mask][cols_export].to_excel(writer, index=False, sheet_name="⚠️ Problèmes")

        # Feuille 3 — Log des corrections barème
        if logs_correction:
            rows = []
            for nom, actions in logs_correction.items():
                for action in actions:
                    rows.append({"Question": nom, "Correction apportée": action})
            pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Corrections barème")

        # Feuille 4 — Résumé
        wb = writer.book
        ws = wb.create_sheet("Résumé")
        total = len(df_analyse)
        nb_err  = (df_analyse["Statut global"] == "❌ Erreur").sum()
        nb_warn = (df_analyse["Statut global"] == "⚠️ Avertissement").sum()
        nb_ok   = (df_analyse["Statut global"] == "✅ OK").sum()
        resume = [
            ("Total questions analysées", total),
            ("✅ Sans problème", nb_ok),
            ("⚠️ Avertissements", nb_warn),
            ("❌ Erreurs critiques", nb_err),
            ("Questions avec barème corrigé", len(logs_correction)),
        ]
        ws["A1"] = "Rapport d'analyse Moodle QA"
        ws["A1"].font = Font(bold=True, size=14)
        for i, (label, val) in enumerate(resume, start=3):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = val

    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
# 7. INTERFACE STREAMLIT
# ═══════════════════════════════════════════════════════════════

# ── Sidebar ───────────────────────────────────────────────────
with st.sidebar:
    st.image("https://upload.wikimedia.org/wikipedia/commons/c/c6/Moodle-logo.svg", width=120)
    st.title("Moodle QA Analyzer")
    st.caption("Analyseur & Correcteur de questions")
    st.divider()

    st.subheader("⚙️ Paramètres")

    penalty_map = {
        "−1/3 (recommandé)": 0.3333333,
        "−1/2": 0.5,
        "−1 (total)": 1.0,
        "0 (sans pénalité)": 0.0,
        "−1/10": 0.1,
    }
    penalty_label = st.selectbox("Pénalité barème cible", list(penalty_map.keys()))
    penalty_cible = penalty_map[penalty_label]

    avec_ia = st.toggle("🤖 Vérification IA (Claude)", value=False,
                        disabled=not bool(ANTHROPIC_API_KEY))
    if not ANTHROPIC_API_KEY:
        st.caption("💡 Ajoutez ANTHROPIC_API_KEY dans les secrets Streamlit pour activer l'IA.")

    corriger_bareme = st.toggle("🔧 Corriger le barème automatiquement", value=True)
    filtre_pb = st.toggle("🔍 Afficher problèmes uniquement", value=False)

    st.divider()
    st.subheader("📊 Filtres")
    type_filter = st.multiselect("Types de questions",
        ["multichoice", "truefalse", "shortanswer", "essay", "numerical", "matching"],
        default=[])

    st.divider()
    st.caption("📚 Formats supportés : .xml .gift")
    st.caption("🔗 [Code source GitHub](#)")

# ── Main ──────────────────────────────────────────────────────
st.title("📚 Analyseur de questions Moodle")
st.markdown("Importez un export Moodle pour **détecter les erreurs** et **corriger le barème** automatiquement.")

# Upload
fichier = st.file_uploader(
    "📂 Choisissez votre fichier Moodle",
    type=["xml", "gift"],
    help="Exportez depuis Moodle → Banque de questions → Exporter"
)

if not fichier:
    col1, col2, col3 = st.columns(3)
    col1.info("**Format XML Moodle**\nExport standard de Moodle, supporte tous les types de questions")
    col2.info("**Format GIFT**\nFormat texte léger, idéal pour les QCM simples")
    col3.info("**Correction automatique**\nAjuste les fractions, pénalités et vérifie le HTML/LaTeX")
    st.stop()

# Chargement
st.info(f"📄 Fichier : **{fichier.name}** ({fichier.size / 1024:.1f} Ko)")
suffix = Path(fichier.name).suffix.lower()

col_btn1, col_btn2 = st.columns([1, 4])
with col_btn1:
    lancer = st.button("🔍 Lancer l'analyse", type="primary", use_container_width=True)

if not lancer:
    st.stop()

# ── Parsing ───────────────────────────────────────────────────
with st.spinner("📥 Chargement des questions..."):
    buf = io.BytesIO(fichier.read())
    if suffix == ".xml":
        questions = charger_xml(buf)
    elif suffix == ".gift":
        questions = charger_gift(buf)
    else:
        st.error("Format non supporté.")
        st.stop()

if not questions:
    st.warning("⚠️ Aucune question détectée. Vérifiez le format du fichier.")
    st.stop()

# Filtre par type
if type_filter:
    questions = [q for q in questions if q["type"] in type_filter]
    if not questions:
        st.warning("Aucune question dans les types sélectionnés.")
        st.stop()

st.success(f"✅ **{len(questions)} question(s)** chargée(s)")

# Numérotation Q01, Q02... ajoutée en préfixe sans modifier le nom original
for i, q in enumerate(questions, start=1):
    q["numero"] = f"Q{i:02d}"
    q["nom_affiche"] = f"Q{i:02d} — {q['nom']}"

# ── Correction barème ─────────────────────────────────────────
logs_correction = {}
questions_corrigees = questions

if corriger_bareme:
    with st.spinner("🔧 Correction du barème en cours..."):
        questions_corrigees, logs_correction = corriger_toutes_questions(questions, penalty_cible)

# ── Analyse ───────────────────────────────────────────────────
progress = st.progress(0, text="Analyse des questions...")
resultats = []

for i, q in enumerate(questions_corrigees):
    res = analyser_question(q)
    if avec_ia and q["type"] == "multichoice":
        res["Vérification IA"] = verifier_ia(q["texte"], q["reponses"], q["bonne"])
    else:
        res["Vérification IA"] = "Non activée" if not avec_ia else "N/A"
    resultats.append(res)
    progress.progress((i + 1) / len(questions_corrigees),
                      text=f"Analyse {i+1}/{len(questions_corrigees)} — {q.get('nom_affiche', q['nom'])[:50]}...")

progress.empty()

df = pd.DataFrame(resultats)

# ── Métriques ─────────────────────────────────────────────────
st.subheader("📊 Résumé")
total   = len(df)
nb_ok   = (df["Statut global"] == "✅ OK").sum()
nb_warn = (df["Statut global"] == "⚠️ Avertissement").sum()
nb_err  = (df["Statut global"] == "❌ Erreur").sum()
nb_corr = len(logs_correction)

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("📋 Total",         total)
c2.metric("✅ Sans problème",  nb_ok,   delta=f"{nb_ok/total*100:.0f}%")
c3.metric("⚠️ Avertissements", nb_warn, delta=f"-{nb_warn}" if nb_warn else "0", delta_color="inverse")
c4.metric("❌ Erreurs",        nb_err,  delta=f"-{nb_err}"  if nb_err  else "0", delta_color="inverse")
c5.metric("🔧 Barèmes corrigés", nb_corr)

# ── Tableau résultats ─────────────────────────────────────────
st.subheader("🔎 Résultats détaillés")

df_affiche = df.copy()
if filtre_pb:
    df_affiche = df_affiche[df_affiche["Statut global"] != "✅ OK"]
    st.caption(f"{len(df_affiche)} question(s) avec problèmes")

cols_show = ["Nom", "Type", "Nb réponses", "Statut global",
             "Texte", "HTML / LaTeX", "Graphique", "Réponses",
             "Encodage", "Barème", "Vérification IA"]
cols_show = [c for c in cols_show if c in df_affiche.columns]

def colorier(val):
    v = str(val)
    if v.startswith("✅"):  return "background-color:#d4edda; color:#155724"
    if v.startswith("⚠️"):  return "background-color:#fff3cd; color:#856404"
    if v.startswith("❌"):  return "background-color:#f8d7da; color:#721c24"
    if v.startswith("🔧"):  return "background-color:#d1ecf1; color:#0c5460"
    return ""

# Compatibilité pandas >= 2.1 : applymap renommé map
try:
    styled = df_affiche[cols_show].style.map(colorier)
except AttributeError:
    styled = df_affiche[cols_show].style.applymap(colorier)

st.dataframe(styled, use_container_width=True, height=420)

# ── Questions problématiques (détail) ────────────────────────
df_pb = df[df["Statut global"] != "✅ OK"].copy()
if not df_pb.empty:
    st.subheader(f"🚨 Questions problématiques ({len(df_pb)})")
    cols_detail = ["Nom", "Type", "Statut global", "HTML / LaTeX", "Graphique",
                   "Réponses", "Encodage", "Barème", "Texte"]
    cols_detail = [c for c in cols_detail if c in df_pb.columns]
    df_err  = df_pb[df_pb["Statut global"] == "❌ Erreur"]
    df_warn = df_pb[df_pb["Statut global"] == "⚠️ Avertissement"]

    if not df_err.empty:
        with st.expander(f"❌ Erreurs critiques — {len(df_err)} question(s)", expanded=True):
            try:
                st.dataframe(df_err[cols_detail].style.map(colorier), use_container_width=True)
            except AttributeError:
                st.dataframe(df_err[cols_detail].style.applymap(colorier), use_container_width=True)
            for _, row in df_err.iterrows():
                st.markdown(f"**🔴 {row['Nom']}** — *{row['Type']}*")
                for col in ["HTML / LaTeX", "Graphique", "Réponses", "Encodage", "Barème", "Texte"]:
                    if col in row and not str(row[col]).startswith("✅"):
                        st.markdown(f"&nbsp;&nbsp;• **{col}** : {row[col]}")
                st.divider()

    if not df_warn.empty:
        with st.expander(f"⚠️ Avertissements — {len(df_warn)} question(s)", expanded=False):
            try:
                st.dataframe(df_warn[cols_detail].style.map(colorier), use_container_width=True)
            except AttributeError:
                st.dataframe(df_warn[cols_detail].style.applymap(colorier), use_container_width=True)
            for _, row in df_warn.iterrows():
                st.markdown(f"**🟡 {row['Nom']}** — *{row['Type']}*")
                for col in ["HTML / LaTeX", "Graphique", "Réponses", "Encodage", "Barème", "Texte"]:
                    if col in row and not str(row[col]).startswith("✅"):
                        st.markdown(f"&nbsp;&nbsp;• **{col}** : {row[col]}")
                st.divider()
else:
    st.success("🎉 Aucune question problématique détectée !")

# ── Log corrections barème ────────────────────────────────────
if logs_correction:
    with st.expander(f"🔧 Détail des corrections barème ({nb_corr} question(s) modifiée(s))"):
        for nom, actions in logs_correction.items():
            st.markdown(f"**{nom}**")
            for a in actions:
                st.markdown(f"  - {a}")

# ── Exports ───────────────────────────────────────────────────
st.subheader("📥 Exports")
col_e1, col_e2 = st.columns(2)

with col_e1:
    excel_buf = generer_excel(df, logs_correction)
    st.download_button(
        "📊 Télécharger rapport Excel",
        data=excel_buf,
        file_name="rapport_moodle_qa.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with col_e2:
    if corriger_bareme and suffix == ".xml":
        xml_corrige = generer_xml_corrige(questions_corrigees)
        st.download_button(
            "📄 Télécharger XML corrigé (Moodle)",
            data=xml_corrige.encode("utf-8"),
            file_name=f"{Path(fichier.name).stem}_corrige.xml",
            mime="application/xml",
            use_container_width=True
        )
    else:
        st.info("Export XML disponible uniquement pour les fichiers .xml avec correction activée.")
